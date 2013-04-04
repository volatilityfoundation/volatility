# Volatility
# Copyright (C) 2008-2011 Volatile Systems
# Copyright (C) 2011 Jamie Levy (Gleeda) <jamie.levy@gmail.com>
#
# This program is free software; you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation; either version 2 of the License, or (at
# your option) any later version.
#
# This program is distributed in the hope that it will be useful, but
# WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the GNU
# General Public License for more details. 
#
# You should have received a copy of the GNU General Public License
# along with this program; if not, write to the Free Software
# Foundation, Inc., 59 Temple Place, Suite 330, Boston, MA 02111-1307 USA 
#

"""
@author:       Jamie Levy (gleeda)
@license:      GNU General Public License 2.0 or later
@contact:      jamie.levy@gmail.com
@organization: Volatile Systems
"""

import volatility.plugins.registry.registryapi as registryapi
import volatility.plugins.taskmods as taskmods
import volatility.plugins.registry.shimcache as shimcache
import volatility.plugins.filescan as filescan
import volatility.plugins.sockets as sockets
import volatility.plugins.sockscan as sockscan
import volatility.plugins.modscan as modscan
import volatility.plugins.procdump as  procdump
import volatility.plugins.dlldump as dlldump
import volatility.plugins.moddump as moddump
import volatility.plugins.netscan as netscan
import volatility.plugins.evtlogs as evtlogs
import volatility.plugins.userassist as userassist
import volatility.plugins.imageinfo as imageinfo
import volatility.win32.rawreg as rawreg
import volatility.addrspace as addrspace
import volatility.win32.tasks as tasks
import volatility.utils as utils
import volatility.protos as protos
import os, sys
import struct
import volatility.debug as debug
import volatility.obj as obj 
import datetime

try:
    from openpyxl.workbook import Workbook
    from openpyxl.writer.excel import ExcelWriter
    from openpyxl.cell import get_column_letter
    has_openpyxl = True 
except ImportError:
    has_openpyxl = False

class TimeLiner(dlldump.DLLDump, procdump.ProcExeDump, evtlogs.EvtLogs, userassist.UserAssist):
    """ Creates a timeline from various artifacts in memory """

    def __init__(self, config, *args):  
        evtlogs.EvtLogs.__init__(self, config, *args)
        config.remove_option("SAVE-EVT")
        userassist.UserAssist.__init__(self, config, *args)
        config.remove_option("HIVE-OFFSET")
        config.remove_option("KEY")
        dlldump.DLLDump.__init__(self, config, *args)
        config.remove_option("BASE")
        config.remove_option("REGEX")
        config.remove_option("IGNORE-CASE")
        procdump.ProcExeDump.__init__(self, config, *args)
        config.remove_option("DUMP-DIR")
        config.remove_option("OFFSET")
        config.remove_option("PID")
        config.remove_option("UNSAFE")

        config.add_option('HIVE', short_option = 'H',
                          help = 'Gather Timestamps from a Particular Registry Hive', type = 'str')
        config.add_option('USER', short_option = 'U',
                          help = 'Gather Timestamps from a Particular User\'s Hive(s)', type = 'str')
        config.add_option("REGISTRY", short_option = "R", default = False, action = 'store_true',
                          help = 'Adds registry keys/dates to timeline')

    def render_text(self, outfd, data):
        for line in data:
            if line != None:
                outfd.write(line)
    
    def render_body(self, outfd, data):
        for line in data:
            if line != None:
                outfd.write(line) 

    def render_xlsx(self, outfd, data):
        wb = Workbook(optimized_write = True)
        ws = wb.create_sheet()
        ws.title = 'Timeline Output'
        for line in data:
            coldata = line.split("|")
            ws.append(coldata)
        wb.save(filename = self._config.OUTPUT_FILE)

    def calculate(self):
        if self._config.OUTPUT == "xlsx" and not has_openpyxl:
            debug.error("You must install OpenPyxl for xlsx format:\n\thttps://bitbucket.org/ericgazoni/openpyxl/wiki/Home")
        elif self._config.OUTPUT == "xlsx" and not self._config.OUTPUT_FILE:
            debug.error("You must specify an output *.xlsx file!\n\t(Example: --output-file=OUTPUT.xlsx)")

        if (self._config.HIVE or self._config.USER) and not (self._config.REGISTRY):
            debug.error("You must use -R/--registry in conjuction with -H/--hive and/or -U/--user")

        addr_space = utils.load_as(self._config)
        version = (addr_space.profile.metadata.get('major', 0), 
                   addr_space.profile.metadata.get('minor', 0))

        pids = {}     #dictionary of process IDs/ImageFileName
        offsets = []  #process offsets
        
        im = imageinfo.ImageInfo(self._config).get_image_time(addr_space) 
        body = False
        if self._config.OUTPUT == "body":
            body = True
    
        if not body:
            event = "{0}|[END LIVE RESPONSE]\n".format(im['ImageDatetime'])
        else:
            event = "0|[END LIVE RESPONSE]|0|---------------|0|0|0|{0}|{0}|{0}|{0}\n".format(im['ImageDatetime'].v())
        yield event
                
        # Get EPROCESS 
        psscan = filescan.PSScan(self._config).calculate()
        for eprocess in psscan:
            if eprocess.obj_offset not in offsets:
                offsets.append(eprocess.obj_offset)

            if not body:
                line = "{0}|{1}|{2}|{3}|{4}|{5}|0x{6:08x}||\n".format(
                    eprocess.CreateTime or '-1', 
                    "[PROCESS]",
                    eprocess.ImageFileName,
                    eprocess.UniqueProcessId,
                    eprocess.InheritedFromUniqueProcessId,
                    eprocess.ExitTime or '',
                    eprocess.obj_offset)
            else:
                line = "0|[PROCESS] {2}/PID: {3}/PPID: {4}/POffset: 0x{5:08x}|0|---------------|0|0|0|{0}|{1}|{0}|{0}\n".format(
                        eprocess.CreateTime.v(), 
                        eprocess.ExitTime.v(),
                        eprocess.ImageFileName,
                        eprocess.UniqueProcessId,
                        eprocess.InheritedFromUniqueProcessId,
                        eprocess.obj_offset)
            pids[eprocess.UniqueProcessId.v()] = eprocess.ImageFileName
            yield line 

        # Get Sockets and Evtlogs XP/2k3 only
        if addr_space.profile.metadata.get('major', 0) == 5:
            socks = sockets.Sockets(self._config).calculate()
            #socks = sockscan.SockScan(self._config).calculate()   # you can use sockscan instead if you uncomment
            for sock in socks:
                la = "{0}:{1}".format(sock.LocalIpAddress, sock.LocalPort)
                if not body:
                    line = "{0}|[SOCKET]|{1}|{2}|Protocol: {3} ({4})|{5:#010x}|||\n".format(
                        sock.CreateTime, 
                        sock.Pid, 
                        la,
                        sock.Protocol,
                        protos.protos.get(sock.Protocol.v(), "-"),
                        sock.obj_offset)
                else:
                    line = "0|[SOCKET] PID: {1}/LocalIP: {2}/Protocol: {3}({4})/POffset: 0x{5:#010x}|0|---------------|0|0|0|{0}|{0}|{0}|{0}\n".format(
                            sock.CreateTime.v(), 
                            sock.Pid,
                            la,
                            sock.Protocol,
                            protos.protos.get(sock.Protocol.v(), "-"),
                            sock.obj_offset)
                yield line

            stuff = evtlogs.EvtLogs.calculate(self)
            for name, buf in stuff:
                for fields in self.parse_evt_info(name, buf, rawtime = True):
                    if not body:
                        line = '{0} |[EVT LOG]|{1}|{2}|{3}|{4}|{5}|{6}|{7}\n'.format(
                            fields[0], fields[1], fields[2], fields[3], fields[4], fields[5], fields[6], fields[7])
                    else:
                        line = "0|[EVT LOG] {1}/{2}/{3}/{4}/{5}/{6}/{7}|0|---------------|0|0|0|{0}|{0}|{0}|{0}\n".format(
                            fields[0].v(),fields[1], fields[2], fields[3], fields[4], fields[5], fields[6], fields[7])
                    yield line
        else:
            # Vista+
            nets = netscan.Netscan(self._config).calculate()
            for net_object, proto, laddr, lport, raddr, rport, state in nets:
                conn = "{0}:{1} -> {2}:{3}".format(laddr, lport, raddr, rport)
                if not body:
                    line = "{0}|[NETWORK CONNECTION]|{1}|{2}|{3}|{4}|{5:<#10x}||\n".format(
                        str(net_object.CreateTime or "-1"),
                        net_object.Owner.UniqueProcessId,
                        conn,
                        proto,
                        state,
                        net_object.obj_offset)
                else:
                    line = "0|[NETWORK CONNECTION] {1}/{2}/{3}/{4}/{5:<#10x}|0|---------------|0|0|0|{0}|{0}|{0}|{0}\n".format(
                        net_object.CreateTime.v(),
                        net_object.Owner.UniqueProcessId,
                        conn,
                        proto,
                        state,
                        net_object.obj_offset)
                yield line

        # Get threads
        threads = modscan.ThrdScan(self._config).calculate()
        for thread in threads:
            image = pids.get(thread.Cid.UniqueProcess.v(), "UNKNOWN")
            if not body:
                line = "{0}|[THREAD]|{1}|{2}|{3}|{4}|||\n".format(
                    thread.CreateTime or '-1',
                    image,
                    thread.Cid.UniqueProcess,
                    thread.Cid.UniqueThread,
                    thread.ExitTime or '',
                    )
            else:
                line = "0|[THREAD] {2}/PID: {3}/TID: {4}|0|---------------|0|0|0|{0}|{1}|{0}|{0}\n".format(
                    thread.CreateTime.v(),
                    thread.ExitTime.v(),
                    image,
                    thread.Cid.UniqueProcess,
                    thread.Cid.UniqueThread,
                    )
            yield line

        # now we get to the PE part.  All PE's are dumped in case you want to inspect them later
    
        data = moddump.ModDump(self._config).calculate()

        for addr_space, procs, mod_base, mod_name in data:
            space = tasks.find_space(addr_space, procs, mod_base)
            if space != None:
                try:
                    header = procdump.ProcExeDump(self._config).get_nt_header(space, mod_base)
                except ValueError, ve: 
                    continue
                try:
                    if not body:
                        line = "{0}|[PE Timestamp (module)]|{1}||{2:#010x}|||||\n".format(
                            header.FileHeader.TimeDateStamp or '-1',
                            mod_name,
                            mod_base)
                    else:
                        line = "0|[PE Timestamp (module)] {1}/Base: {2:#010x}|0|---------------|0|0|0|{0}|{0}|{0}|{0}\n".format(
                            header.FileHeader.TimeDateStamp.v(),
                            mod_name, mod_base)
                except ValueError, ve:
                    if not body:
                        line = "-1|[PE Timestamp (module)]|{0}||{1}|||||\n".format(
                            mod_name,
                            mod_base)
                    else:
                        line = "0|[PE Timestamp (module)] {0}/Base: {1:#010x}|0|---------------|0|0|0|0|0|0|0\n".format(
                            mod_name, mod_base)

                yield line

        # get EPROCESS PE timestamps
        # XXX revert back, now in loop
        for o in offsets:
            self._config.update('OFFSET', o)
            data = self.filter_tasks(procdump.ProcExeDump.calculate(self))
            dllskip = False
            for task in data:
                if task.Peb == None or task.Peb.ImageBaseAddress == None:
                    dllskip = True
                    continue
                try:
                    header = procdump.ProcExeDump(self._config).get_nt_header(task.get_process_address_space(), task.Peb.ImageBaseAddress)
                except ValueError, ve:
                    dllskip = True
                    continue
                try:
                    if not body:
                        line = "{0}|[PE Timestamp (exe)]|{1}|{2}|{3}|{4}|0x{5:08x}|||\n".format(
                            header.FileHeader.TimeDateStamp or "-1",
                            task.ImageFileName,
                            task.UniqueProcessId,
                            task.InheritedFromUniqueProcessId,
                            task.Peb.ProcessParameters.CommandLine,
                            o)
                    else:
                        line = "0|[PE Timestamp (exe)] {1}/PID: {2}/PPID: {3}/Command: {4}/POffset: 0x{5:08x}|0|---------------|0|0|0|{0}|{0}|{0}|{0}\n".format(
                            header.FileHeader.TimeDateStamp.v(),
                            task.ImageFileName,
                            task.UniqueProcessId,
                            task.InheritedFromUniqueProcessId,
                            task.Peb.ProcessParameters.CommandLine,
                            o)

                except ValueError, ve:
                    if not body:
                        line = "-1|[PE Timestamp (exe)]|{0}|{1}|{2}|{3}|0x{4:08x}|||\n".format(
                            task.ImageFileName,
                            task.UniqueProcessId,
                            task.InheritedFromUniqueProcessId,
                            task.Peb.ProcessParameters.CommandLine,
                            o)
                    else:
                        line = "0|[PE Timestamp (exe)] {1}/PID: {2}/PPID: {3}/Command: {4}/POffset: 0x{5:08x}|0|---------------|0|0|0|{0}|{0}|{0}|{0}\n".format(
                            0,
                            task.ImageFileName,
                            task.UniqueProcessId,
                            task.InheritedFromUniqueProcessId,
                            task.Peb.ProcessParameters.CommandLine,
                            o)
                yield line

            # Get DLL PE timestamps
            if not dllskip:
                dlls = self.filter_tasks(dlldump.DLLDump.calculate(self))
            else:
                dllskip = False
                dlls = []
            for proc, ps_ad, base, basename in dlls:
                if ps_ad.is_valid_address(base):
                    if basename == task.ImageFileName:
                        continue
                    try:
                        header = procdump.ProcExeDump(self._config).get_nt_header(ps_ad, base)
                    except ValueError, ve: 
                        continue
                    try:
                        if not body:
                            line = "{0}|[PE Timestamp (dll)]|{1}|{2}|{3}|{4}|EPROCESS Offset: 0x{5:08x}|DLL Base: 0x{6:8x}||\n".format(
                                header.FileHeader.TimeDateStamp or '-1',
                                task.ImageFileName,
                                task.UniqueProcessId,
                                task.InheritedFromUniqueProcessId,
                                basename,
                                o,
                                base)
                        else:
                            line = "0|[PE Timestamp (dll)] {4}/Process: {1}/PID: {2}/PPID: {3}/Process POffset: 0x{5:08x}/DLL Base: 0x{6:8x}|0|---------------|0|0|0|{0}|{0}|{0}|{0}\n".format(
                                header.FileHeader.TimeDateStamp.v(),
                                task.ImageFileName,
                                task.UniqueProcessId,
                                task.InheritedFromUniqueProcessId,
                                basename,
                                o,
                                base)

                    except ValueError, ve:
                        if not body:
                            line = "-1|[PE Timestamp (dll)]|{0}|{1}|{2}|{3}|EPROCESS Offset: 0x{4:08x}|DLL Base: 0x{5:8x}||\n".format(
                                task.ImageFileName,
                                task.UniqueProcessId,
                                task.InheritedFromUniqueProcessId,
                                basename,
                                o,
                                base)
                        else:
                            line = "0|[PE Timestamp (dll)] {4}/Process: {1}/PID: {2}/PPID: {3}/Process POffset: 0x{5:08x}/DLL Base: 0x{6:8x}|0|---------------|0|0|0|{0}|{0}|{0}|{0}\n".format(
                                0,
                                task.ImageFileName,
                                task.UniqueProcessId,
                                task.InheritedFromUniqueProcessId,
                                basename,
                                o,
                                base)
                    yield line

        uastuff = userassist.UserAssist.calculate(self)
        for win7, reg, key in uastuff:
            ts = "{0}".format(key.LastWriteTime)
            for v in rawreg.values(key):
                tp, dat = rawreg.value_data(v)
                subname = v.Name
                if tp == 'REG_BINARY':
                    dat_raw = dat
                    try:
                        subname = subname.encode('rot_13')
                    except UnicodeDecodeError:
                        pass
                    if win7:
                        guid = subname.split("\\")[0]
                        if guid in userassist.folder_guids:
                            subname = subname.replace(guid, userassist.folder_guids[guid])
                    bufferas = addrspace.BufferAddressSpace(self._config, data = dat_raw)
                    uadata = obj.Object("_VOLUSER_ASSIST_TYPES", offset = 0, vm = bufferas)
                    ID = "N/A"
                    count = "N/A"
                    fc = "N/A"
                    tf = "N/A"
                    lw = "N/A"
                    if len(dat_raw) < bufferas.profile.get_obj_size('_VOLUSER_ASSIST_TYPES') or uadata == None:
                        continue
                    else:
                        if hasattr(uadata, "ID"):
                            ID = "{0}".format(uadata.ID)
                        if hasattr(uadata, "Count"):
                            count = "{0}".format(uadata.Count)
                        else:
                            count = "{0}".format(uadata.CountStartingAtFive if uadata.CountStartingAtFive < 5 else uadata.CountStartingAtFive - 5)
                        if hasattr(uadata, "FocusCount"):
                            seconds = (uadata.FocusTime + 500) / 1000.0
                            time = datetime.timedelta(seconds = seconds) if seconds > 0 else uadata.FocusTime
                            fc = "{0}".format(uadata.FocusCount)
                            tf = "{0}".format(time)
                        lw = "{0}".format(uadata.LastUpdated)

                subname = subname.replace("|", "%7c")
                if not body:
                    line = "{0}|[USER ASSIST]|{1}|{2}|{3}|{4}|{5}|{6}\n".format(lw, reg, subname, ID, count, fc, tf)
                else:
                    line = "0|[USER ASSIST] Registry: {1}/Value: {2}/ID: {3}/Count: {4}/FocusCount: {5}/TimeFocused: {6}|0|---------------|0|0|0|{0}|{0}|{0}|{0}\n".format(
                        uadata.LastUpdated.v(), reg, subname, ID, count, fc, tf)
                yield line

        shimdata = shimcache.ShimCache(self._config).calculate()
        for path, lm, lu in shimdata:
            if lu:
                if not body: 
                    line = "{0}|[SHIMCACHE]|{1}|Last update: {2}\n".format(lm, path, lu)
                else:
                    line = "0|[SHIMCACHE] {1}|0|---------------|0|0|0|{0}|{2}|{0}|{0}\n".format(
                        lm.v(), path, lu.v())
            else:
                if not body:
                    line = "{0}|[SHIMCACHE]|{1}|Last update: N/A\n".format(lm, path)
                else:
                    line = "0|[SHIMCACHE] {1}|0|---------------|0|0|0|{0}|{0}|{0}|{0}\n".format(
                        lm.v(), path)
            yield line

        if self._config.REGISTRY:
            regapi = registryapi.RegistryApi(self._config)
            regapi.reset_current()
            regdata = regapi.reg_get_all_keys(self._config.HIVE, self._config.USER, reg = True, rawtime = True)
    
            for lwtime, reg, item in regdata:
                if not body:
                    item = item.replace("|", "%7c")
                    line = "{0:<20}|{1}|{2}\n".format(lwtime, reg, item)
                else:
                    line = "0|[REGISTRY] {1}/{2}|0|---------------|0|0|0|{0}|{0}|{0}|{0}\n".format(
                        lwtime.v(), reg, item)
                yield line

