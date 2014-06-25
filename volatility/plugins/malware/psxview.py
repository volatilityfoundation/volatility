# Volatility
# Copyright (C) 2007-2013 Volatility Foundation
# Copyright (c) 2010, 2011, 2012 Michael Ligh <michael.ligh@mnin.org>
#
# This file is part of Volatility.
#
# Volatility is free software; you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation; either version 2 of the License, or
# (at your option) any later version.
#
# Volatility is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with Volatility.  If not, see <http://www.gnu.org/licenses/>.
#

import volatility.utils as utils
import volatility.obj as obj
import volatility.plugins.common as common
import volatility.win32.tasks as tasks
import volatility.plugins.modscan as modscan
import volatility.plugins.filescan as filescan
import volatility.plugins.overlays.windows.windows as windows
import volatility.plugins.gui.sessions as sessions
import volatility.plugins.gui.windowstations as windowstations

try:
    from openpyxl.workbook import Workbook
    from openpyxl.writer.excel import ExcelWriter
    from openpyxl.cell import get_column_letter
    from openpyxl.style import Color, Fill
    from openpyxl.cell import Cell
    from openpyxl import load_workbook
    has_openpyxl = True 
except ImportError:
    has_openpyxl = False


#--------------------------------------------------------------------------------
# object classes 
#--------------------------------------------------------------------------------

class _PSP_CID_TABLE(windows._HANDLE_TABLE): #pylint: disable-msg=W0212
    """Subclass the Windows handle table object for parsing PspCidTable"""

    def get_item(self, entry, handle_value = 0):

        p = obj.Object("address", entry.Object.v(), self.obj_vm)

        handle = obj.Object("_OBJECT_HEADER",
                offset = (p & ~7) -
                self.obj_vm.profile.get_obj_offset('_OBJECT_HEADER', 'Body'),
                vm = self.obj_vm)

        return handle

#--------------------------------------------------------------------------------
# profile modifications  
#--------------------------------------------------------------------------------

class MalwarePspCid(obj.ProfileModification):
    before = ['WindowsOverlay', 'WindowsVTypes']
    conditions = {'os': lambda x: x == 'windows'}

    def modification(self, profile):
        profile.vtypes.update({"_PSP_CID_TABLE" : profile.vtypes["_HANDLE_TABLE"]})
        profile.merge_overlay({"_KDDEBUGGER_DATA64" : [None,
            {'PspCidTable': [None,
                    ["pointer", ["pointer", ['_PSP_CID_TABLE']]]],
            }]})
        profile.object_classes.update({
            '_PSP_CID_TABLE': _PSP_CID_TABLE,
        })

#--------------------------------------------------------------------------------
# psxview plugin
#--------------------------------------------------------------------------------

class PsXview(common.AbstractWindowsCommand, sessions.SessionsMixin):
    "Find hidden processes with various process listings"

    def __init__(self, config, *args):
        common.AbstractWindowsCommand.__init__(self, config, *args)
        config.add_option("PHYSICAL-OFFSET", short_option = 'P', default = False,
                          help = "Physcal Offset", action = "store_true")
        config.add_option("APPLY-RULES", short_option = 'R', default = False,
                          help = "Apply known good rules", action = "store_true")

    def check_pslist(self, all_tasks):
        """Enumerate processes from PsActiveProcessHead"""
        return dict((p.obj_vm.vtop(p.obj_offset), p) for p in all_tasks)

    def check_psscan(self):
        """Enumerate processes with pool tag scanning"""
        return dict((p.obj_offset, p)
                    for p in filescan.PSScan(self._config).calculate())

    def check_thrdproc(self, _addr_space):
        """Enumerate processes indirectly by ETHREAD scanning"""
        ret = dict()

        for ethread in modscan.ThrdScan(self._config).calculate():
            if ethread.ExitTime != 0:
                continue
            # Bounce back to the threads owner 
            process = None
            if hasattr(ethread.Tcb, 'Process'):
                process = ethread.Tcb.Process.dereference_as('_EPROCESS')
            elif hasattr(ethread, 'ThreadsProcess'):
                process = ethread.ThreadsProcess.dereference()
            # Make sure the bounce succeeded 
            if (process and process.ExitTime == 0 and
                    process.UniqueProcessId > 0 and
                    process.UniqueProcessId < 65535):
                ret[process.obj_vm.vtop(process.obj_offset)] = process

        return ret

    def check_sessions(self, addr_space):
        """Enumerate processes from session structures"""
        
        ret = dict()
        for session in self.session_spaces(addr_space):
            for process in session.processes():
                ret[process.obj_vm.vtop(process.obj_offset)] = process
                
        return ret

    def check_desktop_thread(self, addr_space):
        """Enumerate processes from desktop threads"""
        
        ret = dict()
        for windowstation in windowstations.WndScan(self._config).calculate():
            for desktop in windowstation.desktops():
                for thread in desktop.threads():
                    process = thread.ppi.Process.dereference()
                    if process == None:
                        continue
                    ret[process.obj_vm.vtop(process.obj_offset)] = process
                    
        return ret

    def check_pspcid(self, addr_space):
        """Enumerate processes by walking the PspCidTable"""
        ret = dict()

        # Follow the pointers to the table base
        kdbg = tasks.get_kdbg(addr_space)
        PspCidTable = kdbg.PspCidTable.dereference().dereference()

        # Walk the handle table
        for handle in PspCidTable.handles():
            if handle.get_object_type() == "Process":
                process = handle.dereference_as("_EPROCESS")
                ret[process.obj_vm.vtop(process.obj_offset)] = process

        return ret

    def check_csrss_handles(self, all_tasks):
        """Enumerate processes using the csrss.exe handle table"""
        ret = dict()

        for p in all_tasks:
            if str(p.ImageFileName).lower() == "csrss.exe":
                # Gather the handles to process objects
                for handle in p.ObjectTable.handles():
                    if handle.get_object_type() == "Process":
                        process = handle.dereference_as("_EPROCESS")
                        ret[process.obj_vm.vtop(process.obj_offset)] = process

        return ret

    def calculate(self):
        if self._config.OUTPUT == "xlsx" and not has_openpyxl:
            debug.error("You must install OpenPyxl for xlsx format:\n\thttps://bitbucket.org/ericgazoni/openpyxl/wiki/Home")
        elif self._config.OUTPUT == "xlsx" and not self._config.OUTPUT_FILE:
            debug.error("You must specify an output *.xlsx file!\n\t(Example: --output-file=OUTPUT.xlsx)")

        addr_space = utils.load_as(self._config)

        all_tasks = list(tasks.pslist(addr_space))

        ps_sources = {}
        # The keys are names of process sources. The values
        # are dictionaries whose keys are physical process 
        # offsets and the values are _EPROCESS objects. 
        ps_sources['pslist'] = self.check_pslist(all_tasks)
        ps_sources['psscan'] = self.check_psscan()
        ps_sources['thrdproc'] = self.check_thrdproc(addr_space)
        ps_sources['csrss'] = self.check_csrss_handles(all_tasks)
        ps_sources['pspcid'] = self.check_pspcid(addr_space)
        ps_sources['session'] = self.check_sessions(addr_space)
        if addr_space.profile.metadata.get('major', 0) == 6 and addr_space.profile.metadata.get('minor', 0) >= 2:
            ps_sources['deskthrd'] = {}
        else:
            ps_sources['deskthrd'] = self.check_desktop_thread(addr_space)

        # Build a list of offsets from all sources
        seen_offsets = []
        for source in ps_sources.values():
            for offset in source.keys():
                if offset not in seen_offsets:
                    seen_offsets.append(offset)
                    yield offset, source[offset], ps_sources

    def render_xlsx(self, outfd, data):
        wb = Workbook(optimized_write = True)
        ws = wb.create_sheet()
        ws.title = "Psxview Output"
        ws.append(["Offset (P)",
                  "Name",
                  "PID",
                  "pslist", 
                  "psscan", 
                  "thrdproc", 
                  "pspcid",
                  "csrss", 
                  "session", 
                  "deskthrd",
                  "Exit Time"])
        total = 1
        for offset, process, ps_sources in data:
            incsrss = ps_sources['csrss'].has_key(offset)
            insession = ps_sources['session'].has_key(offset)
            indesktop = ps_sources['deskthrd'].has_key(offset)
            inpspcid = ps_sources['pspcid'].has_key(offset)
            inpslist = ps_sources['pslist'].has_key(offset)
            inthread = ps_sources['thrdproc'].has_key(offset)

            if self._config.APPLY_RULES:
                if not incsrss:
                    if str(process.ImageFileName).lower() in ["system", "smss.exe", "csrss.exe"]:
                        incsrss = "Okay"
                    elif process.ExitTime > 0:
                        incsrss = "Okay"
                if not insession:
                    if str(process.ImageFileName).lower() in ["system", "smss.exe"]:
                        insession = "Okay"
                    elif process.ExitTime > 0:
                        insession = "Okay"
                if not indesktop:
                    if str(process.ImageFileName).lower() in ["system", "smss.exe"]:
                        indesktop = "Okay"
                    elif process.ExitTime > 0:
                        indesktop = "Okay"
                if not inpspcid:
                    if process.ExitTime > 0:
                        inpspcid = "Okay"
                if not inpslist:
                    if process.ExitTime > 0:
                        inpslist = "Okay"
                if not inthread:
                    if process.ExitTime > 0:
                        inthread = "Okay"

            ws.append([hex(offset),
                str(process.ImageFileName),
                str(process.UniqueProcessId),
                str(inpslist),
                str(ps_sources['psscan'].has_key(offset)),
                str(inthread),
                str(inpspcid),
                str(incsrss),
                str(insession),
                str(indesktop),
                str(process.ExitTime or '')])
            total += 1
        wb.save(filename = self._config.OUTPUT_FILE)

        wb = load_workbook(filename = self._config.OUTPUT_FILE)
        ws = wb.get_sheet_by_name(name = "Psxview Output")
        for col in xrange(1, 12):
            ws.cell("{0}{1}".format(get_column_letter(col), 1)).style.font.bold = True
        for row in xrange(2, total + 1):
            for col in xrange(4, 11):
                if ws.cell("{0}{1}".format(get_column_letter(col), row)).value == "False":
                    ws.cell("{0}{1}".format(get_column_letter(col), row)).style.fill.fill_type = Fill.FILL_SOLID
                    ws.cell("{0}{1}".format(get_column_letter(col), row)).style.fill.start_color.index = "FFFF0000"
                else:
                    ws.cell("{0}{1}".format(get_column_letter(col), row)).style.fill.fill_type = Fill.FILL_SOLID
                    ws.cell("{0}{1}".format(get_column_letter(col), row)).style.fill.start_color.index = "FF00FF00" 
        wb.save(filename = self._config.OUTPUT_FILE)

    def render_text(self, outfd, data):

        self.table_header(outfd, [('Offset(P)', '[addrpad]'),
                                  ('Name', '<20'),
                                  ('PID', '>6'),
                                  ('pslist', '5'),
                                  ('psscan', '5'),
                                  ('thrdproc', '5'),
                                  ('pspcid', '5'),
                                  ('csrss', '5'),
                                  ('session', '5'),
                                  ('deskthrd', '5'),
                                  ('ExitTime', ""),
                                  ])

        for offset, process, ps_sources in data:

            incsrss = ps_sources['csrss'].has_key(offset)
            insession = ps_sources['session'].has_key(offset)
            indesktop = ps_sources['deskthrd'].has_key(offset)
            inpspcid = ps_sources['pspcid'].has_key(offset)
            inpslist = ps_sources['pslist'].has_key(offset)
            inthread = ps_sources['thrdproc'].has_key(offset)

            if self._config.APPLY_RULES:
                if not incsrss:
                    if str(process.ImageFileName).lower() in ["system", "smss.exe", "csrss.exe"]:
                        incsrss = "Okay"
                    elif process.ExitTime > 0:
                        incsrss = "Okay"
                if not insession:
                    if str(process.ImageFileName).lower() in ["system", "smss.exe"]:
                        insession = "Okay"
                    elif process.ExitTime > 0:
                        insession = "Okay"
                if not indesktop:
                    if str(process.ImageFileName).lower() in ["system", "smss.exe"]:
                        indesktop = "Okay"
                    elif process.ExitTime > 0:
                        indesktop = "Okay"
                if not inpspcid:
                    if process.ExitTime > 0:
                        inpspcid = "Okay"
                if not inpslist:
                    if process.ExitTime > 0:
                        inpslist = "Okay"
                if not inthread:
                    if process.ExitTime > 0:
                        inthread = "Okay"

            self.table_row(outfd,
                offset,
                process.ImageFileName,
                process.UniqueProcessId,
                str(inpslist),
                str(ps_sources['psscan'].has_key(offset)),
                str(inthread),
                str(inpspcid),
                str(incsrss),
                str(insession),
                str(indesktop),
                str(process.ExitTime or ''), 
                )
