# Volatility
# Copyright (C) 2008-2013 Volatility Foundation
# Copyright (C) 2011 Jamie Levy (Gleeda) <jamie@memoryanalysis.net>
#
# This file is part of Volatility.
#
# Volatility is free software; you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation; either version 2 of the License, or
# (at your option) any later version.
#
# Volatility is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with Volatility.  If not, see <http://www.gnu.org/licenses/>.
#

"""
@author:       Jamie Levy (gleeda)
@license:      GNU General Public License 2.0
@contact:      jamie@memoryanalysis.net
@organization: Volatility Foundation
"""

import volatility.plugins.common as common
import volatility.plugins.registry.registryapi as registryapi
import volatility.plugins.taskmods as taskmods
import volatility.plugins.registry.shimcache as shimcache
import volatility.plugins.filescan as filescan
import volatility.plugins.sockets as sockets
import volatility.plugins.sockscan as sockscan
import volatility.plugins.modscan as modscan
import volatility.plugins.moddump as moddump
import volatility.plugins.netscan as netscan
import volatility.plugins.evtlogs as evtlogs
import volatility.plugins.malware.psxview as psxview
import volatility.plugins.malware.malfind as malfind
import volatility.plugins.registry.userassist as userassist
import volatility.plugins.imageinfo as imageinfo
import volatility.win32.rawreg as rawreg
import volatility.addrspace as addrspace
import volatility.win32.tasks as tasks
import volatility.utils as utils
import volatility.protos as protos
import volatility.plugins.iehistory as iehistory
import os, sys, ntpath
import struct
import volatility.debug as debug
import volatility.obj as obj 
import datetime

try:
    from openpyxl.workbook import Workbook
    from openpyxl.writer.excel import ExcelWriter
    from openpyxl.cell import get_column_letter
    from openpyxl.style import Color, Fill
    from openpyxl.cell import Cell
    from openpyxl import load_workbook
    has_openpyxl = True
except ImportError:
    has_openpyxl = False

# you can have more colors
# http://closedxml.codeplex.com/wikipage?title=Excel%20Indexed%20Colors
colors = {
    "GREEN":   "FF00FF00",
    "LGREEN":  "FFCCFFCC",
    "YELLOW":  "FFFFFF00",
    "RED":     "FFFF0000",
    #"BLUE":    "FF0000FF",
    #"BLUE":    "FF0066CC",
    "BLUE":    "FFCCCCFF",
    "ORANGE":  "FFFF6600",
    "WHITE":   "FFFFFFFF",
    "LORANGE": "FFFF9900",
    "BLUEGRAY": "FF666699",
    "GRAY25":   "FFC0C0C0",
    "TAN":      "FFFFCC99",
    "PINK":     "FFFF00FF",
}
    
class Win7LdrDataTableEntry(obj.ProfileModification):
    before = ['WindowsOverlay']
    conditions = {'os': lambda x: x == 'windows',
                  'major': lambda x: x == 6,
                  'minor': lambda x: x == 1}

    def modification(self, profile):
        overlay = {'_LDR_DATA_TABLE_ENTRY': [ None, {
                        'LoadTime' : [ None, ['WinTimeStamp', dict(is_utc = True)]],
                                        }],
                   # these timestamps need more research for format
                   #'_MMSUPPORT': [ None, {
                   #     'LastTrimStamp': [ None, ['None', dict(is_utc = True)]],
                   #                     }],
                   #'_MMPTE_TIMESTAMP': [ None, {
                   #     'GlobalTimeStamp' : [ None, ['None', dict(is_utc = True)]],
                   #                     }],
                   } 
        profile.merge_overlay(overlay)

class Win7SP1CMHIVE(obj.ProfileModification):
    before = ['WindowsOverlay']
    conditions = {'os': lambda x: x == 'windows',
                  'major': lambda x: x == 6,
                  'minor': lambda x: x == 1,
                  'build': lambda x: x == 7601}

    def modification(self, profile):
        overlay = {'_CMHIVE': [ None, {
                        'LastWriteTime' : [ None, ['WinTimeStamp', dict(is_utc = True)]],
                                        }]}
        profile.merge_overlay(overlay)

class WinXPTrim(obj.ProfileModification):
    before = ['WindowsOverlay']
    conditions = {'os': lambda x: x == 'windows',
                  'major': lambda x: x == 5,
                 }

    def modification(self, profile):
        overlay = {'_MMSUPPORT': [ None, {
                        'LastTrimTime': [ None, ['WinTimeStamp', dict(is_utc = True)]],
                                        }],
                  }
                                        
        profile.merge_overlay(overlay)

class WinAllTime(obj.ProfileModification):
    before = ['WindowsOverlay']
    conditions = {'os': lambda x: x == 'windows',}

    def modification(self, profile):
        overlay = {'_HBASE_BLOCK': [ None, {
                        'TimeStamp' : [ None, ['WinTimeStamp', dict(is_utc = True)]],
                                        }],
                   '_CM_KEY_CONTROL_BLOCK': [ None, {
                        'KcbLastWriteTime': [ None, ['WinTimeStamp', dict(is_utc = True)]],
                                        }],
                   '_IMAGE_DEBUG_DIRECTORY': [ None, {
                        'TimeDateStamp': [ None, ['UnixTimeStamp', dict(is_utc = True)]],
                                        }],
                  } 
        profile.merge_overlay(overlay)

class TimeLiner(common.AbstractWindowsCommand):
    """ Creates a timeline from various artifacts in memory """

    def __init__(self, config, *args, **kwargs):
        common.AbstractWindowsCommand.__init__(self, config, *args, **kwargs)
        config.remove_option("SAVE-EVT")
        config.remove_option("HIVE-OFFSET")
        config.remove_option("KEY")
        config.remove_option("BASE")
        config.remove_option("REGEX")
        config.remove_option("IGNORE-CASE")
        config.remove_option("DUMP-DIR")
        config.remove_option("OFFSET")
        config.remove_option("PID")
        config.remove_option("UNSAFE")

        self.types = ["Process", "Socket", "Shimcache", "Userassist", "IEHistory", "Thread", "Symlink",
                      "_CM_KEY_BODY", "LoadTime", "TimeDateStamp", "_HBASE_BLOCK", "_CMHIVE", "EvtLog", "ImageDate"]

        config.add_option('HIVE', short_option = 'H',
                          help = 'Gather Timestamps from a Particular Registry Hive', type = 'str')
        config.add_option('USER', short_option = 'U',
                          help = 'Gather Timestamps from a Particular User\'s Hive(s)', type = 'str')
        config.add_option('YARA-RULES', short_option = 'Y', default = None,
                        help = 'Yara rules (as a string)')
        config.add_option('YARA-FILE', short_option = 'y', default = None,
                        help = 'Yara rules (rules file)')
        config.add_option("MACHINE", default = "",
                        help = "Machine name to add to timeline header")
        config.add_option("TYPE", default = "".join([",".join(x for x in sorted(self.types))]),
                        help = "Type of artifact to use in timeline (default is all, but \"Registry\")")
        config.add_option("HIGHLIGHT", default = None,
                        help = "Highlight potentially malicious items in xlsx output (experimental)")

        self.suspicious = {}
        self.suspiciouspids = {}

    def render_text(self, outfd, data):
        for line in data:
            if line != None:
                outfd.write(line)
    
    def render_body(self, outfd, data):
        for line in data:
            if line != None:
                outfd.write(line) 

    def fill(self, ws, row, max = 6, color = "RED"):
        for col in xrange(1, max): 
            ws.cell("{0}{1}".format(get_column_letter(col), row)).style.fill.fill_type = Fill.FILL_SOLID
            ws.cell("{0}{1}".format(get_column_letter(col), row)).style.fill.start_color.index = colors.get(color, "RED")

    def render_xlsx(self, outfd, data):
        wb = Workbook(optimized_write = True)
        ws = wb.create_sheet()
        ws.title = 'Timeline Output'
        header = ["Time", "Type", "Item", "Details", "Reason"]
        ws.append(header)
        total = 1
        for line in data:
            coldata = line.split("|")
            ws.append(coldata)
            total += 1
        wb.save(filename = self._config.OUTPUT_FILE)

        if self._config.HIGHLIGHT != None:
            wb = load_workbook(filename = self._config.OUTPUT_FILE)
            ws = wb.get_sheet_by_name(name = "Timeline Output")
            for col in xrange(1, len(header) + 1):
                ws.cell("{0}{1}".format(get_column_letter(col), 1)).style.font.bold = True
            for row in xrange(2, total + 1):
                for col in xrange(2, len(header)):
                    if ws.cell("{0}{1}".format(get_column_letter(col), row)).value in self.suspicious.keys():
                        self.fill(ws, row, len(header) + 1, self.suspicious[ws.cell("{0}{1}".format(get_column_letter(col), row)).value]["color"])
                        ws.cell("{0}{1}".format(get_column_letter(col + 1), row)).value = self.suspicious[ws.cell("{0}{1}".format(get_column_letter(col), row)).value]["reason"]
                    
            wb.save(filename = self._config.OUTPUT_FILE)

    def getoutput(self, header, start, end = None, body = False):
        if body:
            try:
                if end == None:
                    return "0|{0}|0|---------------|0|0|0|{1}|{1}|{1}|{1}\n".format(header, start.v())
                else:
                    return "0|{0}|0|---------------|0|0|0|{1}|{2}|{1}|{1}\n".format(header, start.v(), end.v())
            except ValueError, ve:
                return "0|{0}|0|---------------|0|0|0|{1}|{1}|{1}|{1}\n".format(header, 0)
        else:
            try:
                if end == None or end.v() == 0:
                    return "{0}|{1}\n".format(start, header)
                else:
                    return "{0}|{1} End: {2}\n".format(start, header, end)
            except ValueError, ve:
                return "{0}|{1}\n".format(-1, header)
                
    def calculate(self):
        if self._config.OUTPUT == "xlsx" and not has_openpyxl:
            debug.error("You must install OpenPyxl for xlsx format:\n\thttps://bitbucket.org/ericgazoni/openpyxl/wiki/Home")
        elif self._config.OUTPUT == "xlsx" and not self._config.OUTPUT_FILE:
            debug.error("You must specify an output *.xlsx file!\n\t(Example: --output-file=OUTPUT.xlsx)")

        if (self._config.HIVE or self._config.USER) and "Registry" not in self._config.TYPE:
            debug.error("You must use --registry in conjuction with -H/--hive and/or -U/--user")
        if self._config.TYPE != None:
            for t in self._config.TYPE.split(","):
                if t.strip() not in self.types and t.strip() != "Registry":
                    debug.error("You have entered an incorrect type: {0}".format(t))

        addr_space = utils.load_as(self._config)
        version = (addr_space.profile.metadata.get('major', 0), 
                   addr_space.profile.metadata.get('minor', 0))

        pids = {}     #dictionary of process IDs/ImageFileName
        
        body = False
        if self._config.OUTPUT == "body":
            body = True
        if self._config.MACHINE != "":
            self._config.update("MACHINE", "{0} ".format(self._config.MACHINE))

        detections = False
        if self._config.OUTPUT == "xlsx" and self._config.HIGHLIGHT != None:
            detections = True
    
        if "ImageDate" in self._config.TYPE:
            im = imageinfo.ImageInfo(self._config).get_image_time(addr_space)
            yield self.getoutput("[{0}LIVE RESPONSE]{1} (System time)".format(
                self._config.MACHINE, "" if body else "|"), 
                im['ImageDatetime'], body = body)

        if version <= (6, 1) and "IEHistory" in self._config.TYPE:
            self._config.update("LEAK", True)
            data = iehistory.IEHistory(self._config).calculate()
            for process, record in data:
                ## Extended fields are available for these records 
                if record.obj_name == "_URL_RECORD":
                    line = "[{6}IEHISTORY]{0} {1}->{5}{0} PID: {2}/Cache type \"{3}\" at {4:#x}".format(
                        "" if body else "|",
                        process.ImageFileName,
                        process.UniqueProcessId,
                        record.Signature, record.obj_offset,
                        record.Url,
                        self._config.MACHINE)
                        
                    yield self.getoutput(line, record.LastModified, end = record.LastAccessed, body = body)
            self._config.remove_option("REDR")
            self._config.remove_option("LEAK")

        psx = []
        if "Process" in self._config.Type or "TimeDateStamp" in self._config.Type or \
            "LoadTime" in self._config.Type or "_CM_KEY_BODY" in self._config.Type:
            psx = psxview.PsXview(self._config).calculate()
        for offset, eprocess, ps_sources in psx:
            pids[eprocess.UniqueProcessId.v()] = eprocess.ImageFileName
            if "Process" in self._config.TYPE:
                line = "[{5}PROCESS]{0} {1}{0} PID: {2}/PPID: {3}/POffset: 0x{4:08x}".format(
                    "" if body else "|",
                    eprocess.ImageFileName,
                    eprocess.UniqueProcessId,
                    eprocess.InheritedFromUniqueProcessId,
                    offset,
                    self._config.MACHINE)

                yield self.getoutput(line, eprocess.CreateTime, end = eprocess.ExitTime, body = body)

            if not hasattr(eprocess.obj_vm, "vtop"):
                eprocess = taskmods.DllList(self._config).virtual_process_from_physical_offset(addr_space, eprocess.obj_offset)
                if eprocess == None:
                    continue
            else:
                ps_ad = eprocess.get_process_address_space()
                if ps_ad == None:
                    continue
                
            if version[0] == 5 and "Process" in self._config.TYPE:
                line = "[{5}PROCESS LastTrimTime]{0} {1}{0} PID: {2}/PPID: {3}/POffset: 0x{4:08x}".format(
                    "" if body else "|",
                    eprocess.ImageFileName,
                    eprocess.UniqueProcessId,
                    eprocess.InheritedFromUniqueProcessId,
                    offset,
                    self._config.MACHINE)
                yield self.getoutput(line, eprocess.Vm.LastTrimTime, body = body)

            if eprocess.ObjectTable.HandleTableList and "_CM_KEY_BODY" in self._config.TYPE:
                for handle in eprocess.ObjectTable.handles():
                    if not handle.is_valid():
                        continue

                    name = ""
                    object_type = handle.get_object_type()
                    if object_type == "Key":
                        key_obj = handle.dereference_as("_CM_KEY_BODY")
                        name = key_obj.full_key_name()
                        line = "[{6}Handle (Key)]{0} {1}{0} {2} PID: {3}/PPID: {4}/POffset: 0x{5:08x}".format(
                            "" if body else "|",
                            name,
                            eprocess.ImageFileName,
                            eprocess.UniqueProcessId,
                            eprocess.InheritedFromUniqueProcessId,
                            offset,
                            self._config.MACHINE)
                        yield self.getoutput(line, key_obj.KeyControlBlock.KcbLastWriteTime, body = body)


            if detections and "Process" in self._config.TYPE:
                injected = False
                for vad, address_space in eprocess.get_vads(vad_filter = eprocess._injection_filter):

                    if malfind.Malfind(self._config)._is_vad_empty(vad, address_space):
                        continue

                    line = "PID: {0}/PPID: {1}/POffset: 0x{2:08x}".format(
                        eprocess.UniqueProcessId,
                        eprocess.InheritedFromUniqueProcessId,
                        offset)
                    self.suspicious[line.strip()] = {"reason":"MALFIND", "color": "RED"}
                    self.suspiciouspids[eprocess.UniqueProcessId.v()] = {"reason":"MALFIND", "color": "RED"}
                    injected = True

                proc_okay = False
                if not injected and (eprocess.ExitTime.v() > 0 or str(eprocess.ImageFileName).lower() in ["system", "smss.exe", "csrss.exe"]):
                    proc_okay = True
                elif not injected and ps_sources['psscan'].has_key(offset) and (not ps_sources['pslist'].has_key(offset) or not \
                    ps_sources['thrdproc'].has_key(offset) or not ps_sources['pspcid'].has_key(offset) or not \
                    ps_sources['csrss'].has_key(offset) or not ps_sources['session'].has_key(offset) or not \
                    ps_sources['deskthrd'].has_key(offset)):
                    proc_okay = True
                if not proc_okay and (not ps_sources['pslist'].has_key(offset) or not ps_sources['psscan'].has_key(offset) or not ps_sources['thrdproc'].has_key(offset) \
                    or not ps_sources['pspcid'].has_key(offset) or not ps_sources['csrss'].has_key(offset) or not ps_sources['session'].has_key(offset) \
                    or not ps_sources['deskthrd'].has_key(offset)):
                    if self.suspicious.get(line.strip(), {}).get("reason", None) == None:
                        self.suspicious[line.strip()] = {"reason":"PSXVIEW", "color": "RED"}
                        self.suspiciouspids[eprocess.UniqueProcessId.v()] = {"reason":"PSXVIEW", "color": "RED"}
                    else:
                        self.suspicious[line.strip()] = {"reason":"MALFIND and PSXVIEW", "color": "RED"}
                        self.suspiciouspids[eprocess.UniqueProcessId.v()] = {"reason":"MALFIND and PSXVIEW", "color": "RED"}

            if eprocess.Peb == None or eprocess.Peb.ImageBaseAddress == None:
                continue
            # Get DLL PE timestamps for Wow64 processes (excluding 64-bit ones)
            if eprocess.IsWow64 and "TimeDateStamp" in self._config.TYPE:
                for vad, address_space in eprocess.get_vads(vad_filter = eprocess._mapped_file_filter):
                    if vad.FileObject.FileName:
                        name = str(vad.FileObject.FileName).lower()
                        basename = ntpath.basename(name)
                        if not basename.endswith("dll") or basename in ["wow64cpu.dll", "ntdll.dll", "wow64.dll", "wow64win.dll"]:
                            continue
                        data = ps_ad.zread(vad.Start, vad.Length)
                        bufferas = addrspace.BufferAddressSpace(self._config, data = data)
                        try:
                            pe_file = obj.Object("_IMAGE_DOS_HEADER", offset = 0, vm = bufferas)
                            header = pe_file.get_nt_header()
                        except ValueError, ve: 
                            continue
                        line = "[{7}PE HEADER 32-bit (dll)]{0} {4}{0} Process: {1}/PID: {2}/PPID: {3}/Process POffset: 0x{5:08x}/DLL Base: 0x{6:08x}".format(
                            "" if body else "|",
                            eprocess.ImageFileName,
                            eprocess.UniqueProcessId,
                            eprocess.InheritedFromUniqueProcessId,
                            basename,
                            offset,
                            vad.Start,
                            self._config.MACHINE)
                        yield self.getoutput(line, header.FileHeader.TimeDateStamp, body = body)

            # get DLL PE timestamps
            mods = dict()
            if "TimeDateStamp" in self._config.TYPE or "LoadTime" in self._config.TYPE:
                mods = dict((mod.DllBase.v(), mod) for mod in eprocess.get_load_modules())
            for mod in mods.values():
                basename = str(mod.BaseDllName or "")
                if basename == str(eprocess.ImageFileName):
                    line = "[{7}PE HEADER (exe)]{0} {4}{0} Process: {1}/PID: {2}/PPID: {3}/Process POffset: 0x{5:08x}/DLL Base: 0x{6:08x}".format(
                        "" if body else "|",
                        eprocess.ImageFileName,
                        eprocess.UniqueProcessId,
                        eprocess.InheritedFromUniqueProcessId,
                        basename,
                        offset,
                        mod.DllBase.v(),
                        self._config.MACHINE)
                else:
                    line = "[{7}PE HEADER (dll)]{0} {4}{0} Process: {1}/PID: {2}/PPID: {3}/Process POffset: 0x{5:08x}/DLL Base: 0x{6:08x}".format(
                        "" if body else "|",
                        eprocess.ImageFileName,
                        eprocess.UniqueProcessId,
                        eprocess.InheritedFromUniqueProcessId,
                        basename,
                        offset,
                        mod.DllBase.v(),
                        self._config.MACHINE)
                if "TimeDateStamp" in self._config.TYPE:
                    yield self.getoutput(line, mod.TimeDateStamp, body = body)
                    line2 = "[{7}PE DEBUG]{0} {4}{0} Process: {1}/PID: {2}/PPID: {3}/Process POffset: 0x{5:08x}/DLL Base: 0x{6:08x}".format(
                        "" if body else "|",
                        eprocess.ImageFileName,
                        eprocess.UniqueProcessId,
                        eprocess.InheritedFromUniqueProcessId,
                        basename,
                        offset,
                        mod.DllBase.v(),
                        self._config.MACHINE)
                    yield self.getoutput(line2, mod.get_debug_directory().TimeDateStamp, body = body)
                if hasattr(mod, "LoadTime") and "LoadTime" in self._config.TYPE:
                    temp = line.replace("[{0}PE HEADER ".format(self._config.MACHINE), "[{0}DLL LOADTIME ".format(self._config.MACHINE))
                    yield self.getoutput(temp, mod.TimeDateStamp, end = mod.LoadTime, body = body)
                if detections and eprocess.UniqueProcessId.v() in self.suspiciouspids.keys():
                    suspiciousline = "Process: {0}/PID: {1}/PPID: {2}/Process POffset: 0x{3:08x}/DLL Base: 0x{4:08x}".format(
                        eprocess.ImageFileName,
                        eprocess.UniqueProcessId,
                        eprocess.InheritedFromUniqueProcessId,
                        offset,
                        mod.DllBase.v())
                    self.suspicious[suspiciousline] = {"reason": "Process flagged: " + self.suspiciouspids[eprocess.UniqueProcessId.v()]["reason"], "color": "BLUE"}

        # yarascan
        mal = []
        if self._config.YARA_RULES or self._config.YARA_FILE: 
            mal = malfind.YaraScan(self._config).calculate()
        for o, addr, hit, content in mal:
            # Find out if the hit is from user or kernel mode 
            if o == None:
                continue
            elif o.obj_name == "_EPROCESS":
                line = "PID: {0}/PPID: {1}/POffset: 0x{2:08x}".format(
                    o.UniqueProcessId,
                    o.InheritedFromUniqueProcessId,
                    o.obj_vm.vtop(o.obj_offset))
                if detections and self.suspicious.get(line.strip(), {}).get("reason", None) == None:
                    self.suspicious[line.strip()] = {"reason":"YARASCAN", "color": "LORANGE"}
                    self.suspiciouspids[o.UniqueProcessId.v()] = {"reason":"YARASCAN", "color": "LORANGE"}
                elif self.suspicious[line.strip()]["reason"].find("YARASCAN {0}".format(hit.rule)) == -1:
                    self.suspicious[line.strip()]["reason"] = self.suspicious[line.strip()]["reason"] + " and YARASCAN {0}".format(hit.rule)
            else:
                line = "Base: {0:#010x}".format(
                        o.BaseDllName)
                if detections and self.suspicious.get(line.strip(), {}).get("reason", None) == None:
                    self.suspicious[line.strip()] = {"reason":"YARASCAN {0}".format(hit.rule), "color": "LORANGE"}
                elif detections and self.suspicious[line.strip()]["reason"].find("YARASCAN {0}".format(hit.rule)) == -1:
                    self.suspicious[line.strip()]["reason"] = self.suspicious[line.strip()]["reason"] + " and YARASCAN {0}".format(hit.rule)

        # Get Sockets and Evtlogs XP/2k3 only
        if version[0] == 5:
            #socks = sockets.Sockets(self._config).calculate()
            socks = []
            if "Socket" in self._config.TYPE:
                socks = sockscan.SockScan(self._config).calculate()   # you can use sockscan instead if you uncomment
            for sock in socks:
                la = "{0}:{1}".format(sock.LocalIpAddress, sock.LocalPort)
                line = "[{6}SOCKET]{0} LocalIP: {2}/Protocol: {3}({4}){0} PID: {1}/POffset: 0x{5:#010x}".format(
                        "" if body else "|",
                        sock.Pid, 
                        la, 
                        sock.Protocol,
                        protos.protos.get(sock.Protocol.v(), "-"),
                        sock.obj_offset,
                        self._config.MACHINE)

                yield self.getoutput(line, sock.CreateTime, body = body)
                if detections and sock.Pid.v() in self.suspiciouspids.keys():
                    suspiciousline = "PID: {0}/POffset: 0x{1:#010x}".format(
                        sock.Pid,
                        sock.obj_offset)
                    self.suspicious[suspiciousline] = {"reason": "Process flagged: " + self.suspiciouspids[sock.Pid.v()]["reason"], "color": "YELLOW"}

            stuff = []
            if "EvtLog" in self._config.TYPE:
                evt = evtlogs.EvtLogs(self._config)
                stuff = evt.calculate()
            for name, buf in stuff:
                for fields in evt.parse_evt_info(name, buf, rawtime = True):
                    line = "[{8}EVT LOG]{0} {1}{0} {2}/{3}/{4}/{5}/{6}/{7}".format("" if body else "|",
                            fields[1], fields[2], fields[3], fields[4], fields[5], fields[6], fields[7],
                            self._config.MACHINE)
                    yield self.getoutput(line, fields[0], body = body)
                    if not detections:
                        continue
                    line = "{0}/{1}/{2}/{3}/{4}/{5}".format(
                       fields[2], fields[3], fields[4], fields[5], fields[6], fields[7])
                    if fields[5] == 517:
                        self.suspicious[line] = {"reason": "SEC LOG CLEARED", "color": "RED"}
                    if fields[5] in [861] and fields[6] == "Failure":
                        self.suspicious[line] = {"reason": "PORT BLOCKED", "color": "TAN"}
                    if fields[1].lower() == "secevent.evt" and fields[5] in [529, 680] and fields[6] == "Failure":
                        self.suspicious[line] = {"reason": "LOGIN FAILURE", "color": "TAN"}
                    if fields[1].lower() == "sysevent.evt" and fields[5] in [100]:
                        self.suspicious[line] = {"reason": "LOGIN FAILURE", "color": "TAN"}
                    if fields[1].lower() == "osession.evt" and fields[5] == 7003:
                        self.suspicious[line] = {"reason": "OFFICE APPLICATION FAILURE", "color": "TAN"}
        elif version <= (6, 1):
            # Vista+
            nets = []
            if "Socket" in self._config.TYPE:
                nets = netscan.Netscan(self._config).calculate()
            for net_object, proto, laddr, lport, raddr, rport, state in nets:
                conn = "{0}:{1} -> {2}:{3}".format(laddr, lport, raddr, rport)
                line = "[{6}NETWORK CONNECTION]{0} {2}{0} {1}/{3}/{4}/{5:<#10x}".format(
                        "" if body else "|",
                        net_object.Owner.UniqueProcessId,
                        conn,
                        proto,
                        state,
                        net_object.obj_offset,
                        self._config.MACHINE)

                yield self.getoutput(line, net_object.CreateTime, body = body)
                if detections and net_object.Owner.UniqueProcessId.v() in self.suspiciouspids.keys():
                    suspiciousline = "{0}/{1}/{2}/{3:<#10x}".format(
                        net_object.Owner.UniqueProcessId,
                        proto,
                        state,
                        net_object.obj_offset)
                    self.suspicious[suspiciousline] = {"reason": "Process flagged: " + self.suspiciouspids[net_object.Owner.UniqueProcessId.v()]["reason"], "color": "YELLOW"}

        # Get threads
        threads = []
        if "Thread" in self._config.TYPE:
            threads = modscan.ThrdScan(self._config).calculate()
        for thread in threads:
            image = pids.get(thread.Cid.UniqueProcess.v(), "UNKNOWN")
            line = "[{4}THREAD]{0} {1}{0} PID: {2}/TID: {3}".format(
                    "" if body else "|",
                    image,
                    thread.Cid.UniqueProcess,
                    thread.Cid.UniqueThread,
                    self._config.MACHINE)
            yield self.getoutput(line, thread.CreateTime, end = thread.ExitTime, body = body)

            if not detections:
                continue
            suspiciousline = "PID: {0}/TID: {1}".format(
                    thread.Cid.UniqueProcess,
                    thread.Cid.UniqueThread)
            if thread.Cid.UniqueProcess.v() in self.suspiciouspids.keys():
                self.suspicious[suspiciousline] = {"reason": "Process flagged: " + self.suspiciouspids[thread.Cid.UniqueProcess.v()]["reason"], "color": "YELLOW"}
            if image == "UNKNOWN" or thread.Cid.UniqueProcess.v() not in pids:
                self.suspicious[suspiciousline] = {"reason": "UNKNOWN IMAGE", "color": "YELLOW"}
    
        data = []
        if "Symlink" in self._config.TYPE:
            data = filescan.SymLinkScan(self._config).calculate()
        for link in data:
            objct = link.get_object_header()
            line = "[{6}SYMLINK]{0} {1}->{2}{0} POffset: {3}/Ptr: {4}/Hnd: {5}".format(
                    "" if body else "|",
                    str(objct.NameInfo.Name or ''),
                    str(link.LinkTarget or ''),
                    link.obj_offset,
                    objct.PointerCount,
                    objct.HandleCount,
                    self._config.MACHINE)
            yield self.getoutput(line, link.CreationTime, body = body)

        data = []
        if "TimeDateStamp" in self._config.TYPE:
            data = moddump.ModDump(self._config).calculate()
        for aspace, procs, mod_base, mod_name in data:
            mod_name = str(mod_name or '')
            space = tasks.find_space(aspace, procs, mod_base)
            if space != None:
                try:
                    pe_file = obj.Object("_IMAGE_DOS_HEADER", offset = mod_base, vm = space)
                    header = pe_file.get_nt_header()
                except ValueError, ve: 
                    continue
                line = "[{3}PE HEADER (module)]{0} {1}{0} Base: {2:#010x}".format(
                        "" if body else "|",
                        mod_name,
                        mod_base,
                        self._config.MACHINE)
                yield self.getoutput(line, header.FileHeader.TimeDateStamp, body = body)

        uastuff = []
        if "Userassist" in self._config.TYPE:
            uastuff = userassist.UserAssist(self._config).calculate()
        for win7, reg, key in uastuff:
            ts = "{0}".format(key.LastWriteTime)
            for v in rawreg.values(key):
                tp, dat = rawreg.value_data(v)
                subname = v.Name
                if tp == 'REG_BINARY':
                    dat_raw = dat
                    try:
                        subname = subname.encode('rot_13')
                    except UnicodeDecodeError:
                        pass
                    if win7:
                        guid = subname.split("\\")[0]
                        if guid in userassist.folder_guids:
                            subname = subname.replace(guid, userassist.folder_guids[guid])
                    bufferas = addrspace.BufferAddressSpace(self._config, data = dat_raw)
                    uadata = obj.Object("_VOLUSER_ASSIST_TYPES", offset = 0, vm = bufferas)
                    ID = "N/A"
                    count = "N/A"
                    fc = "N/A"
                    tf = "N/A"
                    lw = "N/A"
                    if len(dat_raw) < bufferas.profile.get_obj_size('_VOLUSER_ASSIST_TYPES') or uadata == None:
                        continue
                    else:
                        if hasattr(uadata, "ID"):
                            ID = "{0}".format(uadata.ID)
                        if hasattr(uadata, "Count"):
                            count = "{0}".format(uadata.Count)
                        else:
                            count = "{0}".format(uadata.CountStartingAtFive if uadata.CountStartingAtFive < 5 else uadata.CountStartingAtFive - 5)
                        if hasattr(uadata, "FocusCount"):
                            seconds = (uadata.FocusTime + 500) / 1000.0
                            time = datetime.timedelta(seconds = seconds) if seconds > 0 else uadata.FocusTime
                            fc = "{0}".format(uadata.FocusCount)
                            tf = "{0}".format(time)
                        lw = "{0}".format(uadata.LastUpdated)

                subname = subname.replace("|", "%7c")
                line = "[{7}USER ASSIST]{0} {2}{0} Registry: {1}/ID: {3}/Count: {4}/FocusCount: {5}/TimeFocused: {6}".format(
                        "" if body else "|",
                        reg, subname, ID, count, fc, tf,
                        self._config.MACHINE)
                yield self.getoutput(line, uadata.LastUpdated, body = body)

        shimdata = []
        if "Shimcache" in self._config.TYPE:
            shimdata = shimcache.ShimCache(self._config).calculate()
        for path, lm, lu in shimdata:
            line = "[{2}SHIMCACHE]{0} {1}{0} ".format(
                    "" if body else "|",
                    path, self._config.MACHINE)
            if lu:
                yield self.getoutput(line, lm, end = lu, body = body)
            else:
                yield self.getoutput(line, lm, body = body)

        if "_HBASE_BLOCK" in self._config.TYPE or "_CMHIVE" in self._config.TYPE or "Registry" in self._config.TYPE:
            regapi = registryapi.RegistryApi(self._config)
            for o in regapi.all_offsets:
                if "_HBASE_BLOCK" in self._config.TYPE:
                    line = "[{2}_HBASE_BLOCK TimeStamp]{0} {1}{0} ".format(
                        "" if body else "|",
                        regapi.all_offsets[o],
                        self._config.MACHINE)
                    h = obj.Object("_HHIVE", o, addr_space)
                    yield self.getoutput(line, h.BaseBlock.TimeStamp, body = body)


                if "_CMHIVE" in self._config.TYPE and version[0] == 6 and addr_space.profile.metadata.get('build', 0) == 7601:
                    line = line = "[{2}_CMHIVE LastWriteTime]{0} {1}{0} ".format(
                        "" if body else "|",
                        regapi.all_offsets[o],
                        self._config.MACHINE)
                    cmhive = obj.Object("_CMHIVE", o, addr_space)
                    yield self.getoutput(line, cmhive.LastWriteTime, body = body)

        if "Registry" in self._config.TYPE:
            regapi.reset_current()
            regdata = regapi.reg_get_all_keys(self._config.HIVE, self._config.USER, reg = True, rawtime = True)
    
            for lwtime, reg, item in regdata:
                item = item.replace("|", "%7c")
                line = "[{3}REGISTRY]{0} {2}{0} Registry: {1}".format(
                        "" if body else "|",
                        reg, 
                        item,
                        self._config.MACHINE)
                        
                yield self.getoutput(line, lwtime, body = body)

