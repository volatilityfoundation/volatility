from volatility import debug
from volatility.renderers.basic import Renderer

__author__ = "gleeda"

try:
    from openpyxl.workbook import Workbook
    from openpyxl.writer.excel import ExcelWriter
    from openpyxl.cell import get_column_letter
    from openpyxl.styles import Color, Fill, Style, PatternFill, Border, Side, Alignment, Protection, Font
    from openpyxl.cell import Cell
    from openpyxl import load_workbook
    has_openpyxl = True 
except ImportError:
    has_openpyxl = False
    
class XLSXRenderer(Renderer):
    def __init__(self, renderers_func, config):
        if not has_openpyxl:
            debug.error("You must install OpenPyxl 2.1.2 for xlsx format:\n\thttps://pypi.python.org/pypi/openpyxl")
        self._config = config
        self._columns = None
        self._text_cell_renderers_func = renderers_func
        self._text_cell_renderers = None
        self._wb = Workbook(optimized_write = True)
        self._ws = self._wb.create_sheet()
        
    def description(self):
        output = []
        for column in self._columns:
            output.append((column.name))
        return output
        
    def _add_row(self, node, data):
        accumulator = data
        accumulator[node] = max(accumulator.values()) + 1
        self._ws.append(list(node.values))
        return accumulator

    def render(self, outfd, grid):
        """Renders the TreeGrid in data out to the output file from the config options"""
        if not self._config.OUTPUT_FILE:
            debug.error("Please specify a valid output file using --output-file")
        self._columns = grid.columns
        self._text_cell_renderers = self._text_cell_renderers_func(self._columns)
        self._ws.append(self.description())
        grid.visit(None, self._add_row, {None: 0}) 
        self._wb.save(filename = self._config.OUTPUT_FILE)
